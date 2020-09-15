from django.shortcuts import render
from rest_framework.authtoken.views import ObtainAuthToken
from rest_framework import viewsets

from .models import (
    TranslationMemory,
    TMContent,
    GlossaryType,
    Glossary,
    GlossaryContent,
    Project,
    File,
    Sentence,
    Corpus,
    CorpusContent,
)

from .serializers import (
    TranslationMemorySerializer,
    TMContentSerializer,
    GlossaryTypeSerializer,
    GlossarySerializer,
    GlossaryContentSerializer,
    ProjectSerializer,
    FileSerializer,
    SentenceSerializer,
    GlossaryWithChildSerializer,
    FileWithChildSerializer,
    ImportTMSerializer,
    ImportGlossarySerializer,
    MachineTranslateSerializer,
    SentenceWithIDSerializer,
    CorpusSerializer,
    CorpusContentSerializer,
    ProjectWithChildSerializer,
)


from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from rest_framework.authtoken.models import Token
from rest_framework.permissions import AllowAny
from rest_framework.views import APIView
from rest_framework.parsers import FileUploadParser

from django.contrib.auth.models import User

from django.http import JsonResponse, HttpResponse, Http404

from elasticsearch import Elasticsearch 
from elasticsearch_dsl import Search, Q

import json
import os

from django.db import IntegrityError
from django.conf import settings
import Levenshtein as lev
import sys
import openpyxl
from rest_framework.parsers import MultiPartParser
from django.db import IntegrityError

from django.core.files import File as _File
sys.path.append(os.path.join(settings.BASE_DIR,'preprocessing_python'))
from preprocessor import *
import requests
import ast
import concurrent.futures
import urllib.request
from googletrans import Translator
from django.db import transaction

import ebooklib
from ebooklib import epub
from bs4 import BeautifulSoup
class TranslationMemoryViewSet(viewsets.ModelViewSet):
    queryset = TranslationMemory.objects.all().order_by('id')
    serializer_class = TranslationMemorySerializer
    
    def get_queryset(self):
        queryset = self.queryset.filter(user_id=self.request.user.id)
        return queryset

class CorpusViewSet(viewsets.ModelViewSet):
    queryset = Corpus.objects.all().order_by('id')
    serializer_class = CorpusSerializer
    
    def get_queryset(self):
        queryset = self.queryset.filter(user_id=self.request.user.id)
        return queryset


class TMContentViewSet(viewsets.ModelViewSet):
    queryset = TMContent.objects.all().order_by('id')
    serializer_class = TMContentSerializer

    def get_queryset(self):
        queryset = self.queryset.filter(translation_memory__user__id = self.request.user.id)
        tm_id = self.request.query_params.get('tm_id', None)
        if tm_id is not None:
            queryset = queryset.filter(translation_memory=tm_id)
        return queryset

class CorpusContentViewSet(viewsets.ModelViewSet):
    queryset = CorpusContent.objects.all().order_by('id')
    serializer_class = CorpusContentSerializer

    def get_queryset(self):
        queryset = self.queryset.filter(corpus__user__id = self.request.user.id)
        corpus_id = self.request.query_params.get('corpus_id', None)
        if corpus_id is not None:
            queryset = queryset.filter(corpus=corpus_id)
        return queryset


class GlossaryTypeViewSet(viewsets.ModelViewSet):
    queryset = GlossaryType.objects.all().order_by('id')
    serializer_class = GlossaryTypeSerializer
    
    def get_queryset(self):
        queryset = self.queryset.filter(user_id=self.request.user.id)
        return queryset


class GlossaryViewSet(viewsets.ModelViewSet):
    queryset = Glossary.objects.all().order_by('id')
    serializer_class = GlossarySerializer
    
    def get_queryset(self):
        queryset = self.queryset.filter(user_id=self.request.user.id)
        return queryset

class GlossaryWithChildViewSet(viewsets.ModelViewSet):
    queryset = Glossary.objects.all().order_by('id')
    serializer_class = GlossaryWithChildSerializer
    
    def get_queryset(self):
        queryset = self.queryset.filter(user_id=self.request.user.id)
        return queryset

class ProjectWithChildViewSet(viewsets.ModelViewSet):
    queryset = Project.objects.all().order_by('id')
    serializer_class = ProjectWithChildSerializer
    
    def get_queryset(self):
        queryset = self.queryset.filter(user_id=self.request.user.id)
        return queryset


class GlossaryContentViewSet(viewsets.ModelViewSet):
    queryset = GlossaryContent.objects.all().order_by('id')
    serializer_class = GlossaryContentSerializer


class ProjectViewSet(viewsets.ModelViewSet):    
    serializer_class = ProjectSerializer
    queryset = Project.objects.all().order_by('id')
    
    def get_queryset(self):
        queryset = self.queryset.filter(user_id=self.request.user.id)
        return queryset


class SentenceViewSet(viewsets.ModelViewSet):
    queryset = Sentence.objects.all().order_by('id')
    serializer_class = SentenceSerializer
    
    def get_queryset(self):
        queryset = self.queryset.filter(file__project__user__id = self.request.user.id)
        file_id = self.request.query_params.get('file_id', None)
        if file_id is not None:
            queryset = queryset.filter(file_id=file_id)
        return queryset


#TODO: fix bug post data to another user
#TODO: fix bug security in upload file (upload to project not belongs to me)
class FileUploadView(APIView):
    def post(self, request, *args, **kwargs):
        from django.utils.datastructures import MultiValueDictKeyError

        try:
            serializer = FileSerializer(data=request.data)
            if "file" not in request.data:
                return Response({"file":["This field is required."]}, status=status.HTTP_400_BAD_REQUEST)          
            if serializer.is_valid():
                text = ''
                # import pdb; pdb.set_trace()
                if request.FILES['file'].name.lower().endswith('.txt'):
                    uf = serializer.save()
                    # Read from file
                    inp_path = os.path.join(settings.BASE_DIR, uf.file.path)
                    with open(inp_path, 'r', encoding='utf8') as f:
                        text = f.read()
                else:    
                    return Response({"detail":"Invalid file type"}, status=status.HTTP_400_BAD_REQUEST)

                # Find src_lang
                project = Project.objects.get(pk=uf.project.id)
                if project.src_lang == settings.VIETNAMESE:
                    p = Preprocessor(Language.vietnamese)    
                elif project.src_lang == settings.ENGLISH:
                    p = Preprocessor(Language.english)    
                else:
                    return Response({"detail":"Invalid Language"}, status=status.HTTP_400_BAD_REQUEST)

                sents = p.segment_to_sentences(text)
                sents_cnt = len(sents)
                
                for idx in range(0, sents_cnt):                
                    sentence_serilizer = SentenceSerializer(data = {"src_str":sents[idx],"file":uf.id})
                    if sentence_serilizer.is_valid():
                        sentence_serilizer.save()

                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except MultiValueDictKeyError:
            return Response({"file":["Invalid file"]}, status=status.HTTP_400_BAD_REQUEST)          
        except ValueError:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def get(self, request, *args, **kwargs):
        try:
            p_id=(self.request.query_params.get('project_id'))
            files = File.objects.filter(project_id=p_id, project__user__id = self.request.user.id)
            serializer = FileSerializer(files, many=True)
            print(serializer.data)
            return Response (serializer.data)
        except ValueError:
            return Response("Please check your input", status=status.HTTP_400_BAD_REQUEST)

class FileUploadDetailView(APIView):
    def get_object(self, pk):
        try:
            return File.objects.get(pk=pk, project__user__id = self.request.user.id)
        except File.DoesNotExist:
            raise Http404

    def get(self, request, pk, format=None):
        file = self.get_object(pk)
        serializer = FileWithChildSerializer(file)
        return Response(serializer.data)       

    def put(self, request, pk, format=None):
        try:
            file = self.get_object(pk)
            # Dont update upload file by remove it from request
            if "file" in request.data:
                del request.data["file"]

            serializer = FileSerializer(file, data=request.data)
            print("request.data = ", request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except ValueError:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


    def delete(self, request, pk, format=None):
        file = self.get_object(pk)
        filePath = file.file.path
        folderPath = os.path.join(settings.DOCUMENT_FOLDER, str(file.project.id))
        # remove soft file
        if os.path.exists(filePath):
            os.remove(filePath)
        # remove empty folder     
        if os.path.exists(folderPath):   
            if len(os.listdir(folderPath)) == 0:
                os.rmdir(folderPath)
        #remove record in db
        file.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

def machine_translation_service(engine, src_lang, tar_lang, sentence):
    if engine == "mymemory":
        url = "https://api.mymemory.translated.net/get?langpair=%s|%s&key=%s&q=%s" % (src_lang, tar_lang, settings.MYMEMORY_KEY,sentence)
        response = requests.get(url)
        byte_str = response.content
        dict_str = byte_str.decode("UTF-8")    
        data = json.loads(dict_str)
        child={}
        child.update({"translation": data["responseData"]["translatedText"]})
        child.update({"source": "MyMemory"})
        return child
    elif engine == "google":
        translator = Translator()
        res = translator.translate(sentence, src=src_lang, dest=tar_lang)
        child={}
        child.update({"translation": res.text})
        child.update({"source": "GoogleTranslate"})
        return child


@api_view(['GET'])
def machine_translate(request):
    try:
        serializer = MachineTranslateSerializer(data=request.data)    
        
        if serializer.is_valid():
            src_lang=serializer.initial_data["src_lang"]
            tar_lang=serializer.initial_data["tar_lang"]
            sentence=serializer.initial_data["sentence"]
            dict=[]
            SERVICES=["mymemory", "google"]

            with concurrent.futures.ThreadPoolExecutor(max_workers = 2) as executor:
                res = {executor.submit(machine_translation_service, e, src_lang, tar_lang, sentence): e for e in SERVICES}
                for future in concurrent.futures.as_completed(res):
                    dict.append(future.result())

            j = {"is_success":True, "err_msg": None} 
            j.update({"result":dict})
            return HttpResponse(json.dumps(j, ensure_ascii=False),
                content_type="application/json",status=status.HTTP_200_OK)
        j = {"is_success":False, "err_msg":  serializer.errors}
        return HttpResponse(json.dumps(j, ensure_ascii=False), content_type="application/json",status=status.HTTP_400_BAD_REQUEST)
    except ValueError as e:
        j = {"is_success":False, "err_msg":  "Failed to Get data: "+str(e)}
        return HttpResponse(json.dumps(j, ensure_ascii=False), content_type="application/json",status=status.HTTP_200_OK)


@api_view(['GET'])
def glossary_find_online_info(request):
    try:
        S = requests.Session()
        url = "https://en.wikipedia.org/w/api.php"
        
        query = request.query_params.get('query', None)
        if query is  None:
            j = {"is_success":False, "err_msg":  "Failed to Search data: "+str(e)}
            return HttpResponse(json.dumps(j, ensure_ascii=False), content_type="application/json",status=status.HTTP_400_BAD_REQUEST)
        
        params = {
            "action": "query",
            "format": "json",
            "list": "search",
            "srlimit": "1",
            "srsearch": query
        }

        R = S.get(url=url, params=params)
        data = R.json()['query']['search']
         
        if data != []:
            dict = {}
            dict["snippet"] =  BeautifulSoup(data[0]['snippet'], 'html5lib').get_text() + "..."
            dict["title"] =  data[0]['title']
            pageid = data[0]['pageid']
            
            params = {
                "action":"query",
                "prop":"info",
                "format": "json",
                "pageids":data[0]['pageid'],
                "inprop":"url"
            }
            R = S.get(url=url, params=params)
            
            data = R.json()

            dict["url"] =  data["query"]['pages'][str(pageid)]['fullurl']

            j = {"is_success":True, "err_msg": None}
            j.update({"result":dict})
        else:
            j = {"is_success":False, "err_msg": "Wiki dont have it"}
        return HttpResponse(json.dumps(j, ensure_ascii=False), content_type="application/json",status=status.HTTP_200_OK)

    except ValueError as e:
        j = {"is_success":False, "err_msg":  "Failed to Search data: "+str(e)}
        return HttpResponse(json.dumps(j, ensure_ascii=False), content_type="application/json",status=status.HTTP_400_BAD_REQUEST)




@api_view(['POST'])
def import_corpus(request):
    try:
        if "file" not in request.data:
            return Response({"file":["This field is required."]}, status=status.HTTP_400_BAD_REQUEST)      
        request_data = request.data
        file_name = request.FILES['file'].name
        request_data["name"] = file_name

        serializer = CorpusSerializer(data=request_data)      
        if serializer.is_valid():
            content = ''
            if file_name.lower().endswith('.epub'):
                book = epub.read_epub(request_data["file"])
                for item in book.get_items_of_type(ebooklib.ITEM_DOCUMENT):
                    soup = BeautifulSoup(item.content, 'html5lib')
                    content = content + soup.get_text()

                content = content.replace('\n', '\r\n')
            elif file_name.lower().endswith('.txt'):
                content = (request_data["file"].read()).decode("utf-8")
            else:
                return Response({"detail":"Invalid file type"}, status=status.HTTP_400_BAD_REQUEST)

            # Find src_lang
            if request_data["language"] == settings.VIETNAMESE:
                p = Preprocessor(Language.vietnamese)    
            elif request_data["language"]  == settings.ENGLISH:
                p = Preprocessor(Language.english)    
            else:
                return Response({"detail":"Invalid Language"}, status=status.HTTP_400_BAD_REQUEST)


            sents = p.segment_to_sentences(content)
            sents_cnt = len(sents)
            
            corpus = serializer.save()
            for idx in range(0, sents_cnt):                
                sentence_refactor = p.preprocess(sents[idx])
                sentence_serilizer = CorpusContentSerializer(data = {"phrase":sentence_refactor,"corpus":corpus.id})
                if sentence_serilizer.is_valid():
                    sentence_serilizer.save()

            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except ValueError:
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



@api_view(['PUT'])
def multi_sentences(request):
    try:
        with transaction.atomic():
            for sentence in request.data:
                sen_obj = Sentence.objects.get(pk=sentence['id'])
                serializer = SentenceSerializer(sen_obj, data=sentence)
                
                if serializer.is_valid():
                    serializer.save()
                else:
                    j = {"is_success":False, "err_msg":  "Failed to Update data: "}
                    return HttpResponse(json.dumps(j, ensure_ascii=False), content_type="application/json",status=status.HTTP_200_OK)
            j = {"is_success":True, "err_msg": None} 
            return HttpResponse(json.dumps(j, ensure_ascii=False),
                content_type="application/json",status=status.HTTP_200_OK)
    except ValueError as e:
        j = {"is_success":False, "err_msg":  "Failed to Update data: "+str(e)}
        return HttpResponse(json.dumps(j, ensure_ascii=False), content_type="application/json",status=status.HTTP_200_OK)


@api_view(['GET'])
def file_download(request):
    try:
        f = File.objects.get(id=request.data["file_id"], project__user__id = request.user.id)
        path = f.file.path
        
        with open(path, 'r') as file :
            filedata = file.read()

        for s in f.sentence_set.all():
            filedata = filedata.replace(s.src_str, s.tar_str, 1)

        new_path = os.path.splitext(path)[0]+" (export)" + os.path.splitext(path)[1]
        
        # with open(new_path, 'w') as file:
        #     file.write(filedata)
        # print(os.path.basename(new_path))

        response = HttpResponse(filedata, content_type="text/plain")
        response['Content-Disposition'] = 'attachment; filename={0}'.format(os.path.basename(new_path))
        return response
    except File.DoesNotExist:
        raise Http404
    except Exception as e:
        j = {"is_success":False, "err_msg":  "Failed to Get data: "+str(e)}
        return HttpResponse(json.dumps(j, ensure_ascii=False), content_type="application/json",status=status.HTTP_400_BAD_REQUEST)

# TODO: write test for api sentence_commit
@api_view(['PUT'])
def sentence_commit(request):
    try:
        tar_str=request.data["tar_str"]
        is_confirmed=request.data["is_confirmed"]
        sentence_id=request.data["sentence_id"]

        user_id = request.user.id

        sentence = Sentence.objects.get(pk=sentence_id, file__project__user_id=user_id)
        sentence.tar_str = tar_str
        sentence.is_confirmed = is_confirmed

        if is_confirmed == True or is_confirmed == 'True':
            if sentence.tm_content:
                tm_content = sentence.tm_content
                tm_content.tar_sentence = tar_str
                tm_content.src_sentence = sentence.src_str
                tm_content.save()
            else:
                tm = sentence.file.project.writable_translation_memory
                tm_content, created = TMContent.objects.get_or_create(translation_memory=tm, tar_sentence = tar_str, src_sentence = sentence.src_str)
                sentence.tm_content=tm_content

            sentence.save()
        else:
            sentence.save()
            tm_content = sentence.tm_content
            tm_content.delete()

        j = {"is_success":True, "err_msg": None} 

        return HttpResponse(json.dumps(j, ensure_ascii=False),
            content_type="application/json",status=status.HTTP_200_OK)

    except Exception as e:
        j = {"is_success":False, "err_msg":  str(e)}
        return HttpResponse(json.dumps(j, ensure_ascii=False), content_type="application/json",status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
def get_tm_by_src_sentence(request):
    try:
        sentence=request.data["sentence"]
        min_similarity=float(request.data["min_similarity"])
        # import pdb; pdb.set_trace()
        client = Elasticsearch([{'host':settings.ELAS_HOST, 'port':settings.ELAS_PORT}])     
        if "translation_memory_id" in request.data: 
            translation_memory_id = request.data["translation_memory_id"]
            q = Q('bool', must=[Q('match', src_sentence=sentence), Q('terms', translation_memory__id=translation_memory_id)])
        else:
            q = Q("match", src_sentence=sentence) 

        s = Search(using=client, index=settings.INDEX_TM).query(q)[0:int(settings.ELAS_NUM_TM_RETURN)] 
        res = s.execute()

        dict=[]
        for i in range(len(res)):
            simi = lev.ratio(sentence, res[i].src_sentence)
            if simi>= min_similarity:
                child={}
                child.update({"src_sentence": res[i].src_sentence})
                child.update({"tar_sentence": res[i].tar_sentence})
                child.update({"translation_memory": res[i].translation_memory.name})
                child.update({"similarity": round(simi, 2)})
                dict.append(child)
        
        dict = sorted(dict, key = lambda i: i['similarity'], reverse=True) 

        j = {"is_success":True, "err_msg": None} 
        j.update({"result":dict})

        return HttpResponse(json.dumps(j, ensure_ascii=False),
            content_type="application/json",status=status.HTTP_200_OK)

    except Exception as e:
        j = {"is_success":False, "err_msg":  "Failed to Get data: "+str(e)}
        return HttpResponse(json.dumps(j, ensure_ascii=False), content_type="application/json",status=status.HTTP_200_OK)


@api_view(['POST'])
def get_glossary_by_src_sentence(request):
    try:
        phrase=request.data["sentence"]
        min_similarity=float(request.data["min_similarity"])

        client = Elasticsearch([{'host':settings.ELAS_HOST, 'port':settings.ELAS_PORT}])   

        if "glossary_id" in request.data: 
            glossary_id = request.data["glossary_id"]
            
            q = Q('bool', must=[Q('match', src_phrase=phrase), Q('terms', glossary__id=glossary_id)])
        else:
            q = Q("match", src_phrase=phrase) 

        s = Search(using=client, index=settings.INDEX_GLOSSARY).query(q)[0:int(settings.ELAS_NUM_GLOSSARY_RETURN)] 
        res = s.execute()

        dict=[]
        for i in range(len(res)):
            simi = lev.ratio(phrase, res[i].src_phrase)
            if simi>= min_similarity:
                if res[i].src_phrase.lower() in phrase.lower():
                    child={}
                    child.update({"src_phrase": res[i].src_phrase})
                    child.update({"tar_phrase": res[i].tar_phrase})
                    child.update({"similarity": round(simi, 2)})
                    child.update({"glossary": res[i].glossary.name})
                    dict.append(child)
        
        dict = sorted(dict, key = lambda i: i['similarity'], reverse=True) 

        j = {"is_success":True, "err_msg": None} 
        j.update({"result":dict})

        return HttpResponse(json.dumps(j, ensure_ascii=False),
            content_type="application/json",status=status.HTTP_200_OK)

    except Exception as e:
        j = {"is_success":False, "err_msg":  "Failed to Get data: "+str(e)}
        return HttpResponse(json.dumps(j, ensure_ascii=False), content_type="application/json",status=status.HTTP_200_OK)

@api_view(['POST'])
def get_corpus_by_phrase(request):
    try:
        phrase=request.data["phrase"]

        client = Elasticsearch([{'host':settings.ELAS_HOST, 'port':settings.ELAS_PORT}])     
        if "user_id" in request.data: 
            user_id = request.data["user_id"]
            q = Q('bool', must=[Q('match_phrase', phrase=phrase), Q('match', corpus__user=user_id)])
        else:
            q = Q("match_phrase", phrase=phrase) 

        s = Search(using=client, index=settings.INDEX_CORPUS).query(q)[0:int(settings.ELAS_NUM_CORPUS_RETURN)] 
        res = s.execute()

        dict=[]
        for i in range(len(res)):
            child={}
            child.update({"phrase": res[i].phrase})
            child.update({"source": res[i].corpus.name})
            dict.append(child)
        
        j = {"is_success":True, "err_msg": None} 
        j.update({"result":dict})

        return HttpResponse(json.dumps(j, ensure_ascii=False),
            content_type="application/json",status=status.HTTP_200_OK)

    except Exception as e:
        j = {"is_success":False, "err_msg":  "Failed to Get data: "+str(e)}
        return HttpResponse(json.dumps(j, ensure_ascii=False), content_type="application/json",status=status.HTTP_200_OK)


# TODO: Add more security
@api_view(['POST'])
@permission_classes((AllowAny,))
def sign_up(request):
    try:
        username = request.data["username"]
        email = request.data["email"]
        password = request.data["password"]

        if User.objects.filter(email=email).exists():
            j = {"is_success":False, "err_msg": email+" already exists"}
            return HttpResponse(json.dumps(j, ensure_ascii=False), content_type="application/json",status=status.HTTP_400_BAD_REQUEST)
        
        user = User.objects.create_user(username, email, password)

        j = {"id":user.id, "username":username, "email":email}

        return HttpResponse(json.dumps(j, ensure_ascii=False), content_type="application/json",status=status.HTTP_200_OK)

    except IntegrityError as e:
        j = {"is_success":False, "err_msg": username+" already exists"}
        return HttpResponse(json.dumps(j, ensure_ascii=False), content_type="application/json",status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        j = {"is_success":False, "err_msg": ""+str(e)}
        return HttpResponse(json.dumps(j, ensure_ascii=False), content_type="application/json",status=status.HTTP_400_BAD_REQUEST)



class ImportGlossaryView(APIView):
    parser_classes = (MultiPartParser,)
    def post(self, request, *args, **kwargs):
        try:
            serializer = ImportGlossarySerializer(data=request.data)    
            glossary_id=request.data["glossary_id"]
            if serializer.is_valid():
                glossary=Glossary.objects.get(pk=glossary_id, user = self.request.user.id)
                # Check invalid file type
                file_name = request.FILES['glossary_file'].name
                if not file_name.lower().endswith(('.xls', '.xlsx')):
                    return Response({"detail":"Invalid file type"}, status=status.HTTP_400_BAD_REQUEST)

                file_obj = request.FILES['glossary_file']

                wb = openpyxl.load_workbook(file_obj)
                ws = wb[wb.sheetnames[0]]

                for row in ws.iter_rows(min_row=2):
                    GlossaryContent.objects.create(
                        src_phrase=row[0].value,
                        tar_phrase=row[1].value,
                        glossary_id=glossary_id
                    )

                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Glossary.DoesNotExist:
            return Response({"detail":"glossary_id not found"}, status=status.HTTP_404_NOT_FOUND)
        except IntegrityError:
            return Response(serializer.data, status=status.HTTP_200_OK)
        except ValueError:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class ImportTMView(APIView):
    parser_classes = (MultiPartParser,)
    def post(self, request, *args, **kwargs):
        try:
            serializer = ImportTMSerializer(data=request.data)    
            tm_id=request.data["tm_id"]
            if serializer.is_valid():
                tm=TranslationMemory.objects.get(pk=tm_id, user = self.request.user.id)
                # Check invalid file type
                file_name = request.FILES['tm_file'].name
                if not file_name.lower().endswith(('.xlsx')):
                    return Response({"detail":"Invalid file type"}, status=status.HTTP_400_BAD_REQUEST)

                file_obj = request.FILES['tm_file']

                wb = openpyxl.load_workbook(file_obj)
                ws = wb[wb.sheetnames[0]]

                for row in ws.iter_rows(min_row=2):
                    TMContent.objects.create(
                        src_sentence=row[0].value,
                        tar_sentence=row[1].value,
                        translation_memory_id=tm_id
                    )
                
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except TranslationMemory.DoesNotExist:
            return Response({"detail":"tm_id not found"}, status=status.HTTP_404_NOT_FOUND)
        except IntegrityError:
            return Response(serializer.data, status=status.HTTP_200_OK)
        except ValueError:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)